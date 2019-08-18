# coding = utf-8
# -*- coding: utf-8 -*-
import json

from openpyxl import load_workbook


# 读取某一个region，从此之下读取两列拼装成map
def read_region_below_map(fn, sheet_name, region_name):
    mp = {}
    wb = load_workbook(fn)
    ws = wb[sheet_name]
    region = ws[region_name]
    next_row = region.row + 1
    column = region.column
    while 1:
        if ws.cell(next_row, column).value is None:
            break
        mp[ws.cell(next_row, column).value] = ws.cell(next_row, column + 1).value
        next_row = next_row + 1
    return mp


def read_sheet_and_get_json(fn, sheet_name, region):
    wb = load_workbook(fn)
    wb.guess_types = True  # 猜测格式类型
    ws = wb[sheet_name]
    region = ws[region]
    row_idx = region.row
    col_idx = region.column
    js_str = ""
    while 1:
        oneline = ws.cell(row_idx, col_idx).value
        if oneline is None:
            break
        oneline = oneline.replace("_x000D_", "")  # 处理excel的换行CR编码残留，略丑陋
        # for i in range(len(oneline)):
        #     print("ascii of " + oneline[i] + " is: " + ascii(ord(oneline[i])))
        js_str = js_str + oneline.strip()
        row_idx = row_idx + 1
    # print(js_str)
    return json.loads(js_str)


# 找到指定的列的关键字keyword所在的行号
def find_row_num(fn, sheet_name, keyword, start_region_name):
    wb = load_workbook(fn)
    ws = wb[sheet_name]
    max_rows = ws.max_row
    start_row = ws[start_region_name].row
    start_col = ws[start_region_name].column
    for i in range(start_row, max_rows):
        candi = ws.cell(i + 1, start_col).value
        if candi == keyword:
            return i + 1
    return -1


# 找到指定的列的关键字keyword所在的坐标
def find_row_region(fn, sheet_name, keyword, start_region_name):
    ws = load_workbook(fn)[sheet_name]
    row_num = find_row_num(fn, sheet_name, keyword, start_region_name)
    nm = ws.cell(row_num, ws[start_region_name].column)
    return nm.coordinate


def find_row_and_pack_map(fn, sheet_name, keyword, start_region_name):
    mp = {}
    ws = load_workbook(fn)[sheet_name]
    row_num = find_row_num(fn, sheet_name, keyword, start_region_name)
    nm = ws.cell(row_num, ws[start_region_name].column)
    region = ws[nm.coordinate]
    next_row = region.row + 1
    column = region.column
    while 1:
        if ws.cell(next_row, column).value is None:
            break
        mp[ws.cell(next_row, column).value] = ws.cell(next_row, column + 1).value
        next_row = next_row + 1
    return mp


class XlsmWrapper:
    def __init__(self, fn):
        self.fn = fn
        self.wb = load_workbook(fn)

    def find_row_and_pack_map(self, sheet_name, keyword, start_region_name):
        mp = {}
        ws = load_workbook(self.fn)[sheet_name]
        row_num = find_row_num(self.fn, sheet_name, keyword, start_region_name)
        nm = ws.cell(row_num, ws[start_region_name].column)
        region = ws[nm.coordinate]
        next_row = region.row + 1
        column = region.column
        while 1:
            if ws.cell(next_row, column).value is None:
                break
            mp[ws.cell(next_row, column).value] = ws.cell(next_row, column + 1).value
            next_row = next_row + 1
        return mp

    def find_row_and_pack_map_with_switch(self, sheet_name, keyword, start_region_name):
        mp = {}
        ws = load_workbook(self.fn)[sheet_name]
        row_num = find_row_num(self.fn, sheet_name, keyword, start_region_name)
        nm = ws.cell(row_num, ws[start_region_name].column)
        region = ws[nm.coordinate]
        next_row = region.row + 1
        column = region.column
        while 1:
            if ws.cell(next_row, column).value is None:
                break
            if ws.cell(next_row, column + 2).value == 'ON':
                mp[ws.cell(next_row, column).value] = ws.cell(next_row, column + 1).value
            next_row = next_row + 1
        return mp
